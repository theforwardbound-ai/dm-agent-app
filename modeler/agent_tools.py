"""ADR 14 — the Investigation Loop's tools and LLM client.

Six read-only, workspace-jailed tools give the model the same
investigative moves GitHub Copilot's harness provides locally: list,
read (line-ranged), search, read Excel sheets, read project state, and
record an unknown. Budgets: per-task iteration cap + total read budget;
when exhausted, tools tell the model to conclude.
"""
import io, json, re, urllib.request
from core import config, state

# ---------------- tool schemas (OpenAI function-calling format) ----------
TOOL_SCHEMAS = [
 {"type": "function", "function": {
   "name": "list_files",
   "description": "List project files (relative path + size). Optionally under a folder prefix like 'input/' or 'output/stage1'.",
   "parameters": {"type": "object", "properties": {
     "prefix": {"type": "string", "description": "folder prefix, default project root"}},
     "required": []}}},
 {"type": "function", "function": {
   "name": "read_file",
   "description": "Read a text file (markdown, sql, yaml, code) by relative path, optionally a line range. Not for xlsx — use read_sheet.",
   "parameters": {"type": "object", "properties": {
     "path": {"type": "string"},
     "start_line": {"type": "integer"}, "end_line": {"type": "integer"}},
     "required": ["path"]}}},
 {"type": "function", "function": {
   "name": "search_files",
   "description": "Search project text files for a regex or phrase; returns path:line matches.",
   "parameters": {"type": "object", "properties": {
     "pattern": {"type": "string"},
     "prefix": {"type": "string"}, "max_hits": {"type": "integer"}},
     "required": ["pattern"]}}},
 {"type": "function", "function": {
   "name": "read_sheet",
   "description": "Read rows from an Excel workbook (e.g. an IDRA). Returns sheet names, headers, and up to max_rows rows as text.",
   "parameters": {"type": "object", "properties": {
     "path": {"type": "string"}, "sheet": {"type": "string"},
     "max_rows": {"type": "integer"}}, "required": ["path"]}}},
 {"type": "function", "function": {
   "name": "get_state",
   "description": "Read project state: 'project' | 'sources' | 'gates' | 'unknowns' | 'defects'.",
   "parameters": {"type": "object", "properties": {
     "section": {"type": "string"}}, "required": ["section"]}}},
 {"type": "function", "function": {
   "name": "record_unknown",
   "description": "Record an open question that inputs cannot answer; it lands in the unknowns register for triage.",
   "parameters": {"type": "object", "properties": {
     "question": {"type": "string"}}, "required": ["question"]}}},
]

class ToolBudget:
    def __init__(self):
        self.read_chars = 0
    def charge(self, text: str) -> str:
        remaining = config.LOOP_READ_BUDGET - self.read_chars
        if remaining <= 0:
            return "[read budget exhausted — conclude now with what you have]"
        clipped = text[:remaining]
        self.read_chars += len(clipped)
        if len(clipped) < len(text):
            clipped += "\n[...truncated: read budget]"
        return clipped

def make_dispatcher(ws, pid: str, source_id: str | None, budget: ToolBudget):
    """Bind tools to one project workspace; every path is jailed by ws."""
    def list_files(prefix: str = ""):
        rows = ws.listdir(prefix)
        return "\n".join(f"{r['path']}  ({r['size']} B)" for r in rows[:400]) \
               or "(empty)"
    def read_file(path: str, start_line: int = 1, end_line: int = 0):
        data = ws.download(path)
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            return "[binary file — use read_sheet for xlsx, or skip]"
        lines = text.splitlines()
        end = end_line if end_line and end_line >= start_line else \
              min(len(lines), start_line + 300 - 1)
        chunk = "\n".join(f"{i}: {l}" for i, l in
                          enumerate(lines[start_line - 1:end], start_line))
        return budget.charge(chunk[:8000])
    def search_files(pattern: str, prefix: str = "", max_hits: int = 30):
        try:
            rx = re.compile(pattern, re.I)
        except re.error:
            rx = re.compile(re.escape(pattern), re.I)
        hits = []
        for r in ws.listdir(prefix):
            if r["path"].endswith((".xlsx", ".png", ".zip")):
                continue
            try:
                text = ws.download(r["path"]).decode("utf-8", errors="ignore")
            except Exception:
                continue
            for i, line in enumerate(text.splitlines(), 1):
                if rx.search(line):
                    hits.append(f"{r['path']}:{i}: {line.strip()[:160]}")
                    if len(hits) >= max_hits:
                        return budget.charge("\n".join(hits))
        return budget.charge("\n".join(hits) or "(no matches)")
    def read_sheet(path: str, sheet: str = "", max_rows: int = 200):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(ws.download(path)), read_only=True,
                           data_only=True)
        names = wb.sheetnames
        # Never silently substitute a different sheet: a wrong guess that
        # quietly returns other data is how a model ends up confidently
        # modelling something it never read.
        if sheet and sheet not in names:
            return (f"[no sheet named '{sheet}' in {path}] "
                    f"available sheets: {', '.join(names)}. "
                    "Call read_sheet again with one of these, or omit sheet "
                    "to read the first.")
        sh = wb[sheet] if sheet else wb[names[0]]
        cap = max(1, min(max_rows, 500))
        body, total = [], 0
        for row in sh.iter_rows(values_only=True):
            total += 1
            if total <= cap:
                body.append(" | ".join("" if c is None else str(c)
                                       for c in row))
        head = [f"sheets: {', '.join(names)}",
                f"reading: {sh.title} ({total} rows)"]
        if total > cap:
            head.append(f"[TRUNCATED: showing {cap} of {total} rows — call "
                        f"read_sheet again with max_rows={min(total, 500)} to "
                        "see the rest before you model anything]")
        return budget.charge("\n".join(head + body))
    def get_state(section: str):
        s = section.lower()
        if s == "project":
            p = state.get_project(pid)
            return json.dumps({k: str(v) for k, v in p.items()}, indent=1)
        if s == "sources":
            return json.dumps(state.sources(pid), indent=1, default=str)
        if s == "gates":
            return "\n".join(sorted(state.confirmed_gates(pid, source_id))) or "(none)"
        if s == "unknowns":
            return json.dumps(state.list_unknowns(pid, source_id), indent=1,
                              default=str)
        if s == "defects":
            return json.dumps(state.list_defects(pid, source_id), indent=1,
                              default=str)
        return "unknown section"
    def record_unknown(question: str):
        state.record_unknown(pid, source_id, question)
        return "recorded"
    return {"list_files": list_files, "read_file": read_file,
            "search_files": search_files, "read_sheet": read_sheet,
            "get_state": get_state, "record_unknown": record_unknown}

def content_text(content) -> str:
    """Normalise assistant content to a string.

    Reasoning models (gpt-oss, and Claude with thinking enabled) return a
    LIST of content blocks rather than a string. Prefer real text blocks;
    fall back to reasoning summaries only if there is no text at all, so an
    iteration never silently returns empty."""
    if content is None:
        return ""
    if isinstance(content, str):
        return content
    if not isinstance(content, list):
        return str(content)
    text, reasoning = [], []
    for b in content:
        if isinstance(b, str):
            text.append(b)
            continue
        if not isinstance(b, dict):
            continue
        kind = b.get("type")
        if kind in ("text", "output_text") and b.get("text"):
            text.append(b["text"])
        elif kind in ("reasoning", "thinking"):
            for s in b.get("summary") or []:
                if isinstance(s, dict) and s.get("text"):
                    reasoning.append(s["text"])
            if b.get("thinking"):
                reasoning.append(b["thinking"])
        elif b.get("text"):
            text.append(b["text"])
    return "\n".join(text) if text else "\n".join(reasoning)


# ---------------- LLM client: three paths ----------------
class LLM:
    """(1) databricks-sdk OpenAI client (on-platform, app SP auth),
    (2) urllib with DM_DATABRICKS_HOST/TOKEN (local real calls),
    (3) fake (offline). Fake supports scripted tool-call sequences so the
    loop's mechanics are testable without a model."""
    def __init__(self, scripted=None):
        self.scripted = list(scripted) if scripted else None
        self._client = None

    def _openai_client(self):
        if self._client is None:
            from databricks.sdk import WorkspaceClient
            self._client = WorkspaceClient().serving_endpoints.get_open_ai_client()
        return self._client

    def chat(self, model: str, messages: list, tools: list | None):
        if self.scripted is not None:
            step = self.scripted.pop(0) if self.scripted else \
                   {"content": "[scripted fake exhausted]"}
            return step, {"in": 0, "out": 0}
        if config.FAKE_LLM:
            return {"content": "[FAKE_LLM] no scripted turns provided"}, \
                   {"in": 0, "out": 0}
        if not config.DBX_HOST:                    # platform path via SDK
            r = self._openai_client().chat.completions.create(
                model=model, messages=messages, tools=tools or None,
                max_tokens=config.MAX_TOKENS_OUT)
            return self._from_openai(r)
        return self._urllib(model, messages, tools)

    def _from_openai(self, r):
        m = r.choices[0].message
        msg = {"content": content_text(m.content)}
        if getattr(m, "tool_calls", None):
            msg["tool_calls"] = [
                {"id": tc.id, "name": tc.function.name,
                 "arguments": tc.function.arguments}
                for tc in m.tool_calls]
        u = getattr(r, "usage", None)
        return msg, {"in": getattr(u, "prompt_tokens", 0) or 0,
                     "out": getattr(u, "completion_tokens", 0) or 0}

    def _urllib(self, model, messages, tools):
        url = f"{config.DBX_HOST.rstrip('/')}/serving-endpoints/{model}/invocations"
        payload = {"messages": messages, "max_tokens": config.MAX_TOKENS_OUT}
        if tools:
            payload["tools"] = tools
        if config.THINKING_BUDGET and "claude" in model:
            payload["thinking"] = {"type": "enabled",
                                   "budget_tokens": config.THINKING_BUDGET}
        req = urllib.request.Request(url, data=json.dumps(payload).encode(),
            headers={"Authorization": f"Bearer {config.DBX_TOKEN}",
                     "Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=600) as resp:
            body = json.loads(resp.read().decode())
        ch = body["choices"][0]["message"]
        msg = {"content": content_text(ch.get("content"))}
        if ch.get("tool_calls"):
            msg["tool_calls"] = [
                {"id": t["id"], "name": t["function"]["name"],
                 "arguments": t["function"]["arguments"]}
                for t in ch["tool_calls"]]
        u = body.get("usage", {})
        return msg, {"in": u.get("prompt_tokens", 0),
                     "out": u.get("completion_tokens", 0)}
