"""Deterministic reconciliation rules over the latest stage-2 bundled
version. Generic set for the Free cycle; office merge tightens against
the bank's templates and rule packs."""
import io, re, yaml
from dataclasses import dataclass

@dataclass
class Defect:
    severity: str
    artifact: str
    description: str

def cols_from_ddl(ddl: str) -> dict[str, set[str]]:
    tables = {}
    for m in re.finditer(r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?"
                         r"[`\"]?([\w.{}`\"]+?)[`\"]?\s*\((.*?)\)\s*"
                         r"(?:COMMENT|;)", ddl, re.S | re.I):
        name = m.group(1).split(".")[-1].strip("`\"").lower()
        cols = set()
        for line in m.group(2).splitlines():
            t = line.strip().strip(",")
            if not t or re.match(r"(?i)^(PRIMARY|FOREIGN|CONSTRAINT|UNIQUE|CHECK)\b", t):
                continue
            cols.add(t.split()[0].strip("`\"").lower())
        tables[name] = cols
    return tables

def cols_from_yaml(text: str) -> dict[str, set[str]]:
    try:
        doc = yaml.safe_load(text)
    except Exception:
        return {}
    out: dict[str, set[str]] = {}
    for t in (doc or {}).get("tables", []):
        name = str(t.get("target_table", "")).lower()
        for c in t.get("columns", []):
            cn = c.get("target_column")
            if name and cn:
                out.setdefault(name, set()).add(str(cn).lower())
    return out

def cols_from_stm(xlsx: bytes) -> dict[str, set[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    out: dict[str, set[str]] = {}
    for sh in wb.worksheets:
        hdr = None
        for row in sh.iter_rows(min_row=1, max_row=5, values_only=True):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if "target column" in cells:
                hdr = cells; break
        if not hdr:
            continue
        ti = hdr.index("target table") if "target table" in hdr else None
        ci = hdr.index("target column")
        pi = hdr.index("privacy") if "privacy" in hdr else None
        for row in sh.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if ci < len(vals) and vals[ci]:
                t = (str(vals[ti]).lower() if ti is not None and ti < len(vals)
                     and vals[ti] else sh.title.lower())
                out.setdefault(t, set()).add(str(vals[ci]).strip().lower())
    return out

def privacy_missing(xlsx: bytes) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    missing = []
    for sh in wb.worksheets:
        hdr = None
        for row in sh.iter_rows(min_row=1, max_row=5, values_only=True):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if "target column" in cells:
                hdr = cells; break
        if not hdr or "privacy" not in hdr:
            continue
        ci, pi = hdr.index("target column"), hdr.index("privacy")
        for row in sh.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if ci < len(vals) and vals[ci] and (pi >= len(vals) or not vals[pi]):
                missing.append(f"{sh.title}.{vals[ci]}")
    return missing

def run_static(artifacts: dict[str, bytes], only: str | None = None) -> list[Defect]:
    ds: list[Defect] = []
    want = lambda a: only in (None, a, "CROSS")
    ddl = next((v.decode(errors="replace") for k, v in artifacts.items()
                if k.endswith(".sql")), None)
    yml = next((v.decode(errors="replace") for k, v in artifacts.items()
                if k.endswith((".yaml", ".yml"))), None)
    stm = next((v for k, v in artifacts.items() if k.endswith(".xlsx")), None)
    md = next((v.decode(errors="replace") for k, v in artifacts.items()
               if k.endswith("-response.md")), "")
    if want("DDL") and not ddl:
        ds.append(Defect("CRITICAL", "DDL", "no DDL artifact in latest version"))
    if want("STM") and not stm:
        ds.append(Defect("CRITICAL", "STM", "no STM workbook in latest version"))
    if want("YAML") and not yml:
        ds.append(Defect("HIGH", "YAML", "no ETL-contract YAML in latest version"))
    if want("STM") and md and "one row per" not in md.lower():
        ds.append(Defect("HIGH", "STM",
                  "no 'One row per ...' grain statement in stage-2 output"))
    if ddl and yml and want("CROSS"):
        d, y = cols_from_ddl(ddl), cols_from_yaml(yml)
        for t in sorted(set(d) & set(y)):
            if d[t] - y[t]:
                ds.append(Defect("HIGH", "CROSS",
                          f"{t}: DDL columns missing from YAML: "
                          f"{sorted(d[t]-y[t])[:6]}"))
            if y[t] - d[t]:
                ds.append(Defect("HIGH", "CROSS",
                          f"{t}: YAML columns missing from DDL: "
                          f"{sorted(y[t]-d[t])[:6]}"))
        for t in sorted(set(y) - set(d)):
            ds.append(Defect("CRITICAL", "CROSS",
                      f"YAML targets table '{t}' absent from DDL"))
    if ddl and stm and want("CROSS"):
        d, s = cols_from_ddl(ddl), cols_from_stm(stm)
        for t in sorted(set(d) & set(s)):
            if s[t] - d[t]:
                ds.append(Defect("CRITICAL", "CROSS",
                          f"{t}: STM maps columns not in DDL: "
                          f"{sorted(s[t]-d[t])[:6]}"))
    if stm and want("STM"):
        miss = privacy_missing(stm)
        if miss:
            ds.append(Defect("HIGH", "STM",
                      f"privacy classification missing for: {miss[:6]}"))
    return ds
